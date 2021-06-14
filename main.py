import socket
import select
import pandas as pd
import openpyxl
import math
import random

# python C:\Users\yyyma\PycharmProjects\testMultipleClientsServerWithValidation\main.py
MAX_MSG_LENGTH = 1024
SERVER_PORT = 5556
SERVER_IP = "0.0.0.0"
COMPETITORS_DATABASE_LOCATION = 'C:/School App Project/competitors database.xlsx'
STUDENTS_DATABASE_LOCATION = 'C:/School App Project/students database.xlsx'


def data_out_students():
    data_in = pd.read_excel(STUDENTS_DATABASE_LOCATION)
    df = pd.DataFrame(data_in, columns=['FirstName', 'LastName', 'AlreadyVoted', 'Class', 'Id', 'PhoneNumber'])
    data_out = df.values.tolist()
    return data_out


def data_out_competitors():
    data_in = pd.read_excel(COMPETITORS_DATABASE_LOCATION)
    df = pd.DataFrame(data_in, columns=['FirstName', 'LastName', 'VotesNumber', 'Class', 'Id', 'Picture', 'CanSee'])
    data_out = df.values.tolist()
    return data_out


def check_student(student_id, student_data):
    for student in student_data:
        if str(student[4]) == student_id:
            if str(student[2]) == "n":
                return "Exist NoVoted"
            else:
                return "Exist Voted"
    return "NoExist NoVoted"


def check_student_phone(student_id, student_phone_number, student_data):
    for student in student_data:
        if str(student[4]) == student_id and "0" + str(student[5]) == student_phone_number:
            if str(student[2]) == "n":
                return "Exist NoVoted"
            else:
                return "Exist Voted"
    return "NoExist NoVoted"


def generateOTP():
    digits = "0123456789"
    OTP = ""
    for i in range(4):
        OTP += digits[math.floor(random.random() * 10)]
    return OTP


def make_onetime_pass(student_id, student_data):
    msg_to_return = ""
    phone_number = ""
    for student in student_data:
        if str(student[4]) == student_id:
            phone_number = "0" + str(student[5])
            if student[2] == "y":
                return "Exist Voted"
    if phone_number == "":
        return "NoExist NoVoted"
    else:
        OTP = generateOTP()
        wbS = openpyxl.load_workbook(STUDENTS_DATABASE_LOCATION)
        sheet = wbS.active
        counter = 2
        while True:
            temp = 'E' + str(counter)
            if str(sheet[temp].value) == str(student_id):
                AlreadyVotedCheckName = 'G' + str(counter)
                sheet[AlreadyVotedCheckName] = OTP
                break
            counter += 1
        wbS.save(STUDENTS_DATABASE_LOCATION)
        msg_to_return = str(phone_number) + " " + str(OTP)
        return msg_to_return


def find_competitors(student_id, student_data, competitor_data):
    competitors_in_class = ""
    student_class = ""
    for student in student_data:
        if str(student[4]) == student_id:
            student_class = student[3]
    for competitor in competitor_data:
        if str(competitor[3]) == student_class:
            competitors_in_class += str(competitor[0]) + " " + str(competitor[1]) + " " + str(competitor[2]) + " " + str(competitor[3]) + " " + str(competitor[4]) + " "
    return competitors_in_class


def get_winner_in_class(class_name, competitor_data):
    winner_in_class = ["", "0"]
    for competitor in competitor_data:
        if str(competitor[3]) == class_name:
            if int(competitor[2]) >= int(winner_in_class[1]):
                winner_in_class[0] = str(competitor[0]) + " " + str(competitor[1])
                winner_in_class[1] = str(competitor[2])
    return winner_in_class


def get_winners(competitor_data):
    to_return = ""
    classes_list = []
    if str(competitor_data[0][6]) == "n":
        return "n"
    else:
        for competitor in competitor_data:
            if str(competitor[3]) not in classes_list:
                classes_list.append(str(competitor[3]))
        for class_name in classes_list:
            winner_in_class = get_winner_in_class(class_name, competitor_data)
            to_return += class_name + " " + winner_in_class[0] + " " + winner_in_class[1] + " "
        print(to_return)
        return to_return


def get_phone_from_id(student_id, student_data):
    for student in student_data:
        if str(student[4]) == student_id:
            if str(student[2]) == "n":
                return "0" + str(student[5])
            else:
                return "Exist Voted"
    return "NoExist NoVoted"


def vote_from_to(student_id, competitor_id, student_data, competitor_data):
    # Set votes number to +1 on the specific competitor
    wbC = openpyxl.load_workbook(COMPETITORS_DATABASE_LOCATION)
    sheet = wbC.active
    counter = 2
    while True:
        temp = 'E' + str(counter)
        if str(sheet[temp].value) == str(competitor_id):
            VoteNumberName = 'C' + str(counter)
            VoteNumberUpdated = int(sheet[VoteNumberName].value) + 1
            sheet[VoteNumberName] = VoteNumberUpdated
            break
        counter += 1
    wbC.save(COMPETITORS_DATABASE_LOCATION)
    # Set alreadyVoted on the specific student
    wbS = openpyxl.load_workbook(STUDENTS_DATABASE_LOCATION)
    sheet = wbS.active
    counter = 2
    while True:
        temp = 'E' + str(counter)
        if str(sheet[temp].value) == str(student_id):
            AlreadyVotedCheckName = 'C' + str(counter)
            sheet[AlreadyVotedCheckName] = 'y'
            break
        counter += 1
    wbS.save(STUDENTS_DATABASE_LOCATION)
    return "good"


# Print students database
competitors_data = data_out_competitors()
print(competitors_data)

# Print competitors database
students_data = data_out_students()
print(students_data)


# Setting up the server
print("Setting up server...")
server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server_socket.bind((SERVER_IP, SERVER_PORT))
server_socket.listen()
print("Listening for clients...")

client_sockets = []
messages_to_send = []


def set_otp(param, param1, param2):
    return "5654"


while True:
    rlist, wlist, xlist = select.select([server_socket] + client_sockets, client_sockets, [])
    for current_socket in rlist:
        if current_socket is server_socket:
            connection, client_address = current_socket.accept()
            print("New client joined!", client_address)
            client_sockets.append(connection)
        else:
            data = current_socket.recv(MAX_MSG_LENGTH).decode()
            if data == "":
                print("Connection closed")
                client_sockets.remove(current_socket)
                current_socket.close()
            else:
                data = str(data[2:])

                data_list = [word for word in data.split(" ")]
                print(data_list)
                answer = ""
                if data_list[0] == "CHECK":
                    answer = check_student(data_list[1], data_out_students())
                elif data_list[0] == "FIND":
                    answer = find_competitors(data_list[1], data_out_students(), data_out_competitors())
                elif data_list[0] == "VOTE":
                    answer = vote_from_to(data_list[1], data_list[2], data_out_students(), data_out_competitors())
                elif data_list[0] == "PHONE":
                    answer = get_phone_from_id(data_list[1], data_out_students())
                elif data_list[0] == "OTP":
                    answer = set_otp(data_list[1], data_list[2], data_out_students())
                elif data_list[0] == "WINNERS":
                    answer = get_winners(data_out_competitors())
                messages_to_send.append((current_socket, answer))

    for message in messages_to_send:
        current_socket, data = message
        if current_socket in wlist:
            if str(data) == "":
                current_socket.send(" ".encode("utf-8"))
                print("massage sent: " + str(data))
                messages_to_send.remove(message)
                current_socket.close()
                client_sockets.remove(current_socket)
                print("Connection closed")
            else:
                current_socket.send(data.encode("utf-8"))
                print("massage sent: " + str(data))
                messages_to_send.remove(message)
                current_socket.close()
                client_sockets.remove(current_socket)
                print("Connection closed")
